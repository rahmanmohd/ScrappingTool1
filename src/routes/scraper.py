from flask import Blueprint, request, jsonify
import requests
from bs4 import BeautifulSoup
import re
import json
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os

scraper_bp = Blueprint('scraper', __name__)

def clean_text(text):
    """Clean and normalize text data."""
    if not text:
        return 'N/A'
    # Remove extra whitespace and normalize
    cleaned = ' '.join(text.split())
    # Remove special characters that might cause issues
    cleaned = re.sub(r'[^\w\s@.-]', '', cleaned)
    return cleaned.strip() if cleaned.strip() else 'N/A'

def extract_emails(text):
    """Extract email addresses from text."""
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    emails = list(set(re.findall(email_pattern, text)))
    # Filter out common false positives
    filtered_emails = []
    for email in emails:
        if not any(skip in email.lower() for skip in ['example.com', 'test.com', 'placeholder']):
            filtered_emails.append(email)
    return filtered_emails

def extract_phones(text):
    """Extract phone numbers from text."""
    phone_patterns = [
        r'(\+?1?[-.\s]?)?(\(?[0-9]{3}\)?[-.\s]?[0-9]{3}[-.\s]?[0-9]{4})',
        r'(\+?[0-9]{1,3}[-.\s]?)?(\(?[0-9]{3,4}\)?[-.\s]?[0-9]{3,4}[-.\s]?[0-9]{3,4})'
    ]
    
    phones = []
    for pattern in phone_patterns:
        matches = re.findall(pattern, text)
        for match in matches:
            phone = ''.join(match).strip()
            if len(phone) >= 10:  # Minimum phone number length
                phones.append(phone)
    
    return list(set(phones))

def extract_company_info(soup):
    """Extract additional company information."""
    info = {}
    
    # Try to extract company description from various sources
    description_sources = [
        soup.find('meta', attrs={'name': 'description'}),
        soup.find('meta', attrs={'property': 'og:description'}),
        soup.find('meta', attrs={'name': 'twitter:description'})
    ]
    
    for source in description_sources:
        if source and source.get('content'):
            info['description'] = clean_text(source['content'])
            break
    else:
        info['description'] = 'N/A'
    
    # Extract industry/keywords from meta tags
    keywords_meta = soup.find('meta', attrs={'name': 'keywords'})
    if keywords_meta and keywords_meta.get('content'):
        info['keywords'] = clean_text(keywords_meta['content'])
    else:
        info['keywords'] = 'N/A'
    
    # Try to find company address
    address_patterns = [
        r'\d+\s+[\w\s]+(?:Street|St|Avenue|Ave|Road|Rd|Drive|Dr|Lane|Ln|Boulevard|Blvd)',
        r'\d+\s+[\w\s]+,\s*[\w\s]+,\s*[A-Z]{2}\s*\d{5}'
    ]
    
    page_text = soup.get_text()
    for pattern in address_patterns:
        matches = re.findall(pattern, page_text, re.IGNORECASE)
        if matches:
            info['address'] = clean_text(matches[0])
            break
    else:
        info['address'] = 'N/A'
    
    return info

def scrape_company_data(url):
    """Scrape company data from a given URL."""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Extract company name
        company_name = 'N/A'
        if soup.title:
            company_name = clean_text(soup.title.string)
        
        # Try to get a cleaner company name from h1 tags or specific selectors
        name_selectors = ['h1', '.company-name', '#company-name', '.brand-name']
        for selector in name_selectors:
            elements = soup.select(selector)
            if elements and elements[0].text.strip():
                company_name = clean_text(elements[0].text)
                break
        
        # Extract website URL
        website_url = url
        canonical = soup.find('link', rel='canonical')
        if canonical and canonical.get('href'):
            website_url = canonical['href']
        
        # Extract all text for contact info search
        page_text = soup.get_text()
        
        # Extract contact information
        emails = extract_emails(page_text)
        phones = extract_phones(page_text)
        
        # Extract additional company info
        additional_info = extract_company_info(soup)
        
        # Extract social media links
        social_links = []
        social_domains = ['linkedin.com', 'twitter.com', 'facebook.com', 'instagram.com', 'youtube.com']
        for link in soup.find_all('a', href=True):
            href = link['href']
            for domain in social_domains:
                if domain in href and href not in social_links:
                    social_links.append(href)
                    break
        
        return {
            'company_name': company_name,
            'website_url': website_url,
            'emails': emails[:5],  # Limit to first 5 emails
            'phones': phones[:3],  # Limit to first 3 phone numbers
            'description': additional_info.get('description', 'N/A'),
            'keywords': additional_info.get('keywords', 'N/A'),
            'address': additional_info.get('address', 'N/A'),
            'social_links': social_links[:5],  # Limit to first 5 social links
            'status': 'success',
            'scraped_at': str(pd.Timestamp.now()) if 'pd' in globals() else 'N/A'
        }
        
    except Exception as e:
        return {
            'company_name': 'N/A',
            'website_url': url,
            'emails': [],
            'phones': [],
            'description': 'N/A',
            'keywords': 'N/A',
            'address': 'N/A',
            'social_links': [],
            'status': 'error',
            'error': str(e),
            'scraped_at': str(pd.Timestamp.now()) if 'pd' in globals() else 'N/A'
        }

@scraper_bp.route('/scrape', methods=['POST'])
def scrape_urls():
    """Scrape multiple URLs and return company data."""
    data = request.get_json()
    urls = data.get('urls', [])
    
    if not urls:
        return jsonify({'error': 'No URLs provided'}), 400
    
    results = []
    for url in urls:
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        
        result = scrape_company_data(url)
        results.append(result)
    
    return jsonify({'results': results})

@scraper_bp.route('/export', methods=['POST'])
def export_data():
    """Export scraped data in various formats."""
    data = request.get_json()
    results = data.get('results', [])
    format_type = data.get('format', 'json')
    
    if not results:
        return jsonify({'error': 'No data to export'}), 400
    
    if format_type == 'json':
        return jsonify(results)
    
    elif format_type == 'csv':
        output = io.StringIO()
        fieldnames = ['company_name', 'website_url', 'emails', 'phones', 'description', 
                     'keywords', 'address', 'social_links', 'status', 'scraped_at']
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for result in results:
            # Convert lists to strings for CSV
            csv_result = result.copy()
            csv_result['emails'] = '; '.join(result.get('emails', []))
            csv_result['phones'] = '; '.join(result.get('phones', []))
            csv_result['social_links'] = '; '.join(result.get('social_links', []))
            writer.writerow(csv_result)
        
        csv_content = output.getvalue()
        output.close()
        
        return csv_content, 200, {
            'Content-Type': 'text/csv',
            'Content-Disposition': 'attachment; filename=scraped_data.csv'
        }
    
    elif format_type == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Scraped Company Data"
        
        # Define headers
        headers = ['Company Name', 'Website URL', 'Emails', 'Phone Numbers', 
                  'Description', 'Keywords', 'Address', 'Social Links', 'Status', 'Scraped At']
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for row, result in enumerate(results, 2):
            ws.cell(row=row, column=1, value=result.get('company_name', 'N/A'))
            ws.cell(row=row, column=2, value=result.get('website_url', 'N/A'))
            ws.cell(row=row, column=3, value='; '.join(result.get('emails', [])))
            ws.cell(row=row, column=4, value='; '.join(result.get('phones', [])))
            ws.cell(row=row, column=5, value=result.get('description', 'N/A'))
            ws.cell(row=row, column=6, value=result.get('keywords', 'N/A'))
            ws.cell(row=row, column=7, value=result.get('address', 'N/A'))
            ws.cell(row=row, column=8, value='; '.join(result.get('social_links', [])))
            ws.cell(row=row, column=9, value=result.get('status', 'N/A'))
            ws.cell(row=row, column=10, value=result.get('scraped_at', 'N/A'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Read file content
        with open(temp_file.name, 'rb') as f:
            excel_content = f.read()
        
        # Clean up temporary file
        os.unlink(temp_file.name)
        
        return excel_content, 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': 'attachment; filename=scraped_data.xlsx'
        }
    
    else:
        return jsonify({'error': 'Unsupported format'}), 400

@scraper_bp.route('/analyze', methods=['POST'])
def analyze_data():
    """Analyze scraped data and provide insights."""
    data = request.get_json()
    results = data.get('results', [])
    
    if not results:
        return jsonify({'error': 'No data to analyze'}), 400
    
    analysis = {
        'total_companies': len(results),
        'successful_scrapes': len([r for r in results if r.get('status') == 'success']),
        'failed_scrapes': len([r for r in results if r.get('status') == 'error']),
        'companies_with_emails': len([r for r in results if r.get('emails')]),
        'companies_with_phones': len([r for r in results if r.get('phones')]),
        'companies_with_social': len([r for r in results if r.get('social_links')]),
        'companies_with_address': len([r for r in results if r.get('address', 'N/A') != 'N/A']),
        'success_rate': 0
    }
    
    if analysis['total_companies'] > 0:
        analysis['success_rate'] = round((analysis['successful_scrapes'] / analysis['total_companies']) * 100, 2)
    
    # Extract most common domains
    domains = {}
    for result in results:
        if result.get('website_url'):
            try:
                from urllib.parse import urlparse
                domain = urlparse(result['website_url']).netloc
                domains[domain] = domains.get(domain, 0) + 1
            except:
                pass
    
    analysis['top_domains'] = sorted(domains.items(), key=lambda x: x[1], reverse=True)[:5]
    
    return jsonify(analysis)

